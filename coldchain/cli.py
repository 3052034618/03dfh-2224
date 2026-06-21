import csv
import os
import sys
from collections import defaultdict
from datetime import datetime
from typing import Optional, List

import click

from .loader import load_waybills, load_temperature_dir, find_temperature_files, find_waybill_files
from .matcher import match_waybills_readings
from .models import TempZone, WaybillReport, ExceptionItem
from .reporter import generate_report
from .analyzer import detect_exceptions


_SEVERITY_LABEL = {"high": "严重", "medium": "中等", "low": "轻微"}
_CATEGORY_LABEL = {
    "gap": "数据断点",
    "no_data": "无数据",
    "spike": "温度突升/降",
    "over_temp": "超温",
}


@click.group()
@click.version_option("1.2.0", prog_name="coldchain")
def cli():
    """冷链运单温度留痕工具 —— 批量整理温度记录，生成留痕报告与异常清单"""
    pass


@cli.command()
@click.option("--waybill-file", "-w", required=True, help="运单清单文件路径（CSV/XLSX）")
@click.option("--temp-dir", "-t", required=True, help="温度记录文件夹路径")
@click.option("--output", "-o", default=None, help="匹配结果输出文件路径（CSV/XLSX）")
def match(waybill_file, temp_dir, output):
    """校验匹配：按运单号/车牌/设备编号匹配，提示缺失项"""
    if not os.path.exists(waybill_file):
        click.echo(f"错误：运单文件不存在 → {waybill_file}", err=True)
        sys.exit(1)
    if not os.path.isdir(temp_dir):
        click.echo(f"错误：温度记录文件夹不存在 → {temp_dir}", err=True)
        sys.exit(1)

    click.echo("正在加载运单清单...")
    waybills = load_waybills(waybill_file)
    click.echo(f"  已加载 {len(waybills)} 条运单")

    click.echo("正在加载温度记录...")
    readings = load_temperature_dir(temp_dir)
    click.echo(f"  已加载 {len(readings)} 条温度记录")

    click.echo("正在匹配...")
    result = match_waybills_readings(waybills, readings)

    click.echo()
    click.echo("=" * 60)
    click.echo("匹配结果摘要")
    click.echo("=" * 60)
    click.echo(f"  匹配成功：{len(result.matched)} 票")
    click.echo(f"  无温度数据：{len(result.no_temp_data)} 票")
    click.echo(f"  缺少起运时间：{len(result.no_departure)} 票")
    click.echo(f"  缺少到达时间：{len(result.no_arrival)} 票")
    click.echo(f"  无设备编号：{len(result.no_device_id)} 票")
    click.echo(f"  未匹配设备：{len(result.unmatched_devices)} 个")

    if result.matched:
        click.echo()
        click.echo(click.style("✓ 匹配成功的运单：", fg="green"))
        for pair in result.matched:
            w = pair.waybill
            click.echo(f"  - {w.waybill_no}  车牌:{w.license_plate or '—'}  "
                       f"设备:{w.device_id or '—'}  "
                       f"匹配依据:{pair.basis}  "
                       f"记录数:{len(pair.readings)}")

    if result.no_departure:
        click.echo()
        click.echo(click.style("⚠ 缺少起运时间的运单：", fg="yellow"))
        for w in result.no_departure:
            click.echo(f"  - {w.waybill_no}  车牌:{w.license_plate or '—'}  设备:{w.device_id or '—'}")

    if result.no_arrival:
        click.echo()
        click.echo(click.style("⚠ 缺少到达时间的运单：", fg="yellow"))
        for w in result.no_arrival:
            click.echo(f"  - {w.waybill_no}  车牌:{w.license_plate or '—'}  设备:{w.device_id or '—'}")

    if result.no_temp_data:
        click.echo()
        click.echo(click.style("✗ 无对应温度数据的运单：", fg="red"))
        for w in result.no_temp_data:
            click.echo(f"  - {w.waybill_no}  车牌:{w.license_plate or '—'}  设备:{w.device_id or '—'}")

    if result.no_device_id:
        click.echo()
        click.echo(click.style("⚠ 无设备编号的运单（无法匹配温度数据）：", fg="yellow"))
        for w in result.no_device_id:
            click.echo(f"  - {w.waybill_no}  车牌:{w.license_plate or '—'}")

    if result.unmatched_devices:
        click.echo()
        click.echo(click.style("ℹ 以下设备有温度数据但未匹配到运单：", fg="cyan"))
        for dev in result.unmatched_devices:
            click.echo(f"  - {dev}")

    if output:
        _ensure_dir(output)
        if output.lower().endswith((".xlsx", ".xls")):
            _write_excel(output, _build_match_rows(result), sheet_name="匹配结果")
        else:
            _write_match_csv(output, result)
        click.echo()
        click.echo(f"匹配结果已导出 → {output}")


def _build_match_rows(result):
    rows = []
    for pair in result.matched:
        w = pair.waybill
        rows.append({
            "运单号": w.waybill_no, "车牌": w.license_plate, "设备编号": w.device_id,
            "客户": w.customer, "温区": w.raw_temp_zone,
            "起运时间": _fmt_dt(w.departure_time), "到达时间": _fmt_dt(w.arrival_time),
            "匹配状态": "已匹配", "匹配依据": pair.basis, "温度记录条数": len(pair.readings),
        })
    for w in result.no_temp_data:
        rows.append({
            "运单号": w.waybill_no, "车牌": w.license_plate, "设备编号": w.device_id,
            "客户": w.customer, "温区": w.raw_temp_zone,
            "起运时间": _fmt_dt(w.departure_time), "到达时间": _fmt_dt(w.arrival_time),
            "匹配状态": "无温度数据", "匹配依据": "", "温度记录条数": 0,
        })
    return rows


def _write_match_csv(filepath, result):
    _ensure_dir(filepath)
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["运单号", "车牌", "设备编号", "客户", "温区", "起运时间", "到达时间",
                         "匹配状态", "匹配依据", "温度记录条数"])
        for row in _build_match_rows(result):
            writer.writerow([row["运单号"], row["车牌"], row["设备编号"], row["客户"],
                             row["温区"], row["起运时间"], row["到达时间"],
                             row["匹配状态"], row["匹配依据"], row["温度记录条数"]])


@cli.command()
@click.option("--waybill-file", "-w", required=True, help="运单清单文件路径（CSV/XLSX）")
@click.option("--temp-dir", "-t", required=True, help="温度记录文件夹路径")
@click.option("--customer", "-c", default=None, help="筛选客户名称（模糊匹配）")
@click.option("--temp-zone", "-z", default=None, help="覆盖温区标准（如 -18~-15 或 2-8）")
@click.option("--output", "-o", default=None, help="报告输出文件路径（CSV/XLSX）")
def report(waybill_file, temp_dir, customer, temp_zone, output):
    """留痕报告：输出每票运单的最低温、最高温、超温分钟数和关键时间点"""
    if not os.path.exists(waybill_file):
        click.echo(f"错误：运单文件不存在 → {waybill_file}", err=True)
        sys.exit(1)
    if not os.path.isdir(temp_dir):
        click.echo(f"错误：温度记录文件夹不存在 → {temp_dir}", err=True)
        sys.exit(1)

    override_zone = None
    if temp_zone:
        override_zone = TempZone.parse(temp_zone)
        if not override_zone:
            click.echo(f'错误：无法解析温区标准「{temp_zone}」，请检查格式', err=True)
            sys.exit(1)

    waybills = load_waybills(waybill_file)
    readings = load_temperature_dir(temp_dir)

    if customer:
        waybills = [w for w in waybills if customer in (w.customer or "")]

    match_result = match_waybills_readings(waybills, readings)

    click.echo()
    click.echo("=" * 70)
    click.echo("冷链运单温度留痕报告")
    click.echo("=" * 70)
    if customer:
        click.echo(f"客户筛选：{customer}")
    if override_zone:
        click.echo(click.style(f"温区标准（命令行指定）：{override_zone.display()}", bold=True))
    elif waybills and waybills[0].temp_zone:
        click.echo("温区标准：采用运单清单中的温区")

    reports = []
    all_exceptions = []
    for pair in match_result.matched:
        zone = override_zone or pair.waybill.temp_zone
        r = generate_report(pair.waybill, pair.readings, zone, pair.basis)
        reports.append(r)
        _print_waybill_report(r)
        items = detect_exceptions(pair.waybill, pair.readings, zone)
        all_exceptions.extend(items)

    for w in match_result.no_temp_data:
        zone = override_zone or w.temp_zone
        r = generate_report(w, [], override_zone)
        reports.append(r)
        _print_waybill_report(r)
        items = detect_exceptions(w, [], zone)
        all_exceptions.extend(items)

    click.echo()
    click.echo(f"共 {len(reports)} 票运单"
               f"（有数据 {sum(1 for r in reports if r.has_data)} 票，"
               f"无数据 {sum(1 for r in reports if not r.has_data)} 票）")

    if output:
        _ensure_dir(output)
        if output.lower().endswith((".xlsx", ".xls")):
            _export_full_excel(output, match_result, reports, override_zone,
                               all_exceptions=all_exceptions)
        else:
            _write_report_csv(output, reports)
        click.echo(f"报告已导出 → {output}")
        if output.lower().endswith((".xlsx", ".xls")) and all_exceptions:
            click.echo(f"  （包含异常清单页，共 {len(all_exceptions)} 条异常）")


def _print_waybill_report(r):
    click.echo()
    click.echo(click.style(f"▶ 运单 {r.waybill_no}", bold=True))
    click.echo(f"  车牌：{r.license_plate or '—'}    设备：{r.device_id or '—'}    客户：{r.customer or '—'}")
    if r.temp_zone:
        click.echo(f"  温区标准：{r.temp_zone.display()}")
    if r.match_basis:
        click.echo(f"  匹配依据：{r.match_basis}")
    click.echo(f"  起运：{_fmt_dt(r.departure_time)}    到达：{_fmt_dt(r.arrival_time)}")

    if not r.has_data:
        click.echo(click.style("  ✗ 无温度数据", fg="red"))
        return

    click.echo(f"  最低温：{r.min_temp}℃    最高温：{r.max_temp}℃    记录条数：{r.reading_count}")
    if r.temp_zone:
        over_label = click.style(f"{r.over_temp_minutes} 分钟", fg="red") if r.over_temp_minutes > 0 else "0 分钟"
        click.echo(f"  超温时长（按{r.temp_zone.display()}标准）：{over_label}")

        if r.over_temp_segments:
            click.echo("  超温时段：")
            for seg in r.over_temp_segments:
                click.echo(f"    {_fmt_dt(seg.start)} ~ {_fmt_dt(seg.end)}  "
                           f"({seg.minutes}分钟)  {seg.min_temp}℃~{seg.max_temp}℃")

    click.echo("  关键时间点：")
    for kp in r.key_time_points:
        click.echo(f"    {_fmt_dt(kp.timestamp)}  {kp.temperature}℃  [{kp.label}]")


def _build_report_rows(reports):
    rows = []
    for r in reports:
        rows.append({
            "运单号": r.waybill_no, "车牌": r.license_plate, "设备编号": r.device_id,
            "客户": r.customer,
            "温区标准": r.temp_zone.display() if r.temp_zone else "",
            "匹配依据": r.match_basis,
            "起运时间": _fmt_dt(r.departure_time), "到达时间": _fmt_dt(r.arrival_time),
            "最低温(℃)": r.min_temp if r.min_temp is not None else "",
            "最高温(℃)": r.max_temp if r.max_temp is not None else "",
            "记录条数": r.reading_count,
            "超温分钟数": r.over_temp_minutes,
            "有数据": "是" if r.has_data else "否",
        })
    return rows


def _build_keypoint_rows(reports):
    rows = []
    for r in reports:
        for kp in r.key_time_points:
            rows.append({
                "运单号": r.waybill_no,
                "时间": _fmt_dt(kp.timestamp),
                "温度(℃)": kp.temperature,
                "标签": kp.label,
            })
    return rows


def _write_report_csv(filepath, reports):
    _ensure_dir(filepath)
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        headers = ["运单号", "车牌", "设备编号", "客户", "温区标准", "匹配依据",
                   "起运时间", "到达时间", "最低温(℃)", "最高温(℃)",
                   "记录条数", "超温分钟数", "有数据"]
        writer.writerow(headers)
        for row in _build_report_rows(reports):
            writer.writerow([row[h] for h in headers])

        writer.writerow([])
        writer.writerow(["=== 关键时间点明细 ==="])
        writer.writerow(["运单号", "时间", "温度(℃)", "标签"])
        for row in _build_keypoint_rows(reports):
            writer.writerow([row["运单号"], row["时间"], row["温度(℃)"], row["标签"]])


@cli.command()
@click.option("--waybill-file", "-w", required=True, help="运单清单文件路径（CSV/XLSX）")
@click.option("--temp-dir", "-t", required=True, help="温度记录文件夹路径")
@click.option("--temp-zone", "-z", default=None, help="覆盖温区标准（如 -18~-15 或 2-8）")
@click.option("--severity", "-s", type=click.Choice(["high", "medium", "low", "all"]),
              default="all", help="筛选严重程度")
@click.option("--output", "-o", default=None, help="异常清单输出文件路径（CSV/XLSX）")
@click.option("--note-template", "-n", is_flag=True, default=False,
              help="导出时增加备注列、处理人列和处理状态列")
@click.option("--merge-notes", "-m", "merge_notes_file", default=None,
              help="已有备注文件路径（CSV/XLSX），导出时保留原说明")
def exceptions(waybill_file, temp_dir, temp_zone, severity, output, note_template, merge_notes_file):
    """异常清单：按严重程度列出断点、无数据、温度突升等情况"""
    if not os.path.exists(waybill_file):
        click.echo(f"错误：运单文件不存在 → {waybill_file}", err=True)
        sys.exit(1)
    if not os.path.isdir(temp_dir):
        click.echo(f"错误：温度记录文件夹不存在 → {temp_dir}", err=True)
        sys.exit(1)

    override_zone = None
    if temp_zone:
        override_zone = TempZone.parse(temp_zone)
        if not override_zone:
            click.echo(f'错误：无法解析温区标准「{temp_zone}」，请检查格式', err=True)
            sys.exit(1)
        click.echo(click.style(f"温区标准（命令行指定）：{override_zone.display()}", bold=True))

    waybills = load_waybills(waybill_file)
    readings = load_temperature_dir(temp_dir)

    match_result = match_waybills_readings(waybills, readings)

    existing_notes = {}
    if merge_notes_file and note_template:
        click.echo(f"正在合并已有备注 → {merge_notes_file}")
        existing_notes = _load_existing_notes(merge_notes_file)
        if existing_notes:
            click.echo(f"  已读取 {len(existing_notes)} 条历史备注")
    elif merge_notes_file and not note_template:
        click.echo(click.style("  提示：--merge-notes 需要配合 -n 使用才会输出备注列", fg="yellow"))

    all_items = []
    for pair in match_result.matched:
        zone = override_zone or pair.waybill.temp_zone
        items = detect_exceptions(pair.waybill, pair.readings, zone)
        all_items.extend(items)

    for w in match_result.no_temp_data:
        zone = override_zone or w.temp_zone
        items = detect_exceptions(w, [], zone)
        all_items.extend(items)

    if severity != "all":
        all_items = [it for it in all_items if it.severity == severity]

    all_items.sort(key=lambda x: x.sort_key())

    click.echo()
    click.echo("=" * 70)
    click.echo("冷链温度异常清单")
    click.echo("=" * 70)

    if not all_items:
        click.echo(click.style("  未检测到异常 ✓", fg="green"))
        return

    high_count = sum(1 for it in all_items if it.severity == "high")
    medium_count = sum(1 for it in all_items if it.severity == "medium")
    low_count = sum(1 for it in all_items if it.severity == "low")
    click.echo(f"  严重 {high_count}  |  中等 {medium_count}  |  轻微 {low_count}")
    click.echo()

    for idx, it in enumerate(all_items, 1):
        sev_label = _SEVERITY_LABEL.get(it.severity, it.severity)
        cat_label = _CATEGORY_LABEL.get(it.category, it.category)
        color = {"high": "red", "medium": "yellow", "low": "white"}.get(it.severity, "white")

        header = f"#{idx}  [{sev_label}]  {cat_label}"
        click.echo(click.style(f"  {header}", fg=color))
        click.echo(f"  运单：{it.waybill_no}  车牌：{it.license_plate or '—'}  设备：{it.device_id or '—'}")
        if it.start_time and it.end_time:
            click.echo(f"  时间：{_fmt_dt(it.start_time)} ~ {_fmt_dt(it.end_time)}"
                       f"（{it.minutes}分钟）")
        if it.temperature is not None:
            click.echo(f"  温度：{it.temperature}℃")
        click.echo(f"  详情：{it.detail}")
        click.echo()

    if output:
        _ensure_dir(output)
        if output.lower().endswith((".xlsx", ".xls")):
            _write_excel(output, _build_exception_rows(all_items, note_template, existing_notes),
                         sheet_name="异常清单", highlight_severity=True)
        else:
            _write_exceptions_csv(output, all_items, note_template, existing_notes)
        click.echo(f"异常清单已导出 → {output}")
        if note_template:
            click.echo(f"  （含备注模板：处理状态下拉、备注、处理人）")


def _build_exception_rows(items, note_template=False, existing_notes=None):
    if existing_notes is None:
        existing_notes = {}
    rows = []
    for idx, it in enumerate(items, 1):
        existing = existing_notes.get(it.unique_key(), {})
        row = {
            "序号": idx,
            "严重程度": _SEVERITY_LABEL.get(it.severity, it.severity),
            "类别": _CATEGORY_LABEL.get(it.category, it.category),
            "运单号": it.waybill_no,
            "车牌": it.license_plate,
            "设备编号": it.device_id,
            "开始时间": _fmt_dt(it.start_time),
            "结束时间": _fmt_dt(it.end_time),
            "时长(分钟)": it.minutes,
            "温度(℃)": it.temperature if it.temperature is not None else "",
            "详情": it.detail,
        }
        if note_template:
            row["处理状态"] = existing.get("处理状态", it.status or "")
            row["备注"] = existing.get("备注", it.remark or "")
            row["处理人"] = existing.get("处理人", it.handler or "")
        rows.append(row)
    return rows


def _load_existing_notes(filepath):
    if not os.path.exists(filepath):
        return {}
    notes = {}
    ext = os.path.splitext(filepath)[1].lower()
    try:
        def _normalize_time(text):
            text = str(text).strip()
            if text in ("—", "-", "", "None"):
                return ""
            return text

        def _normalize_mins(text):
            text = str(text).strip()
            if not text:
                return ""
            try:
                return f"{float(text):.1f}"
            except ValueError:
                return text

        def _normalize_temp(text):
            text = str(text).strip()
            if not text:
                return ""
            try:
                val = float(text)
                if val == int(val):
                    return str(int(val))
                return str(val)
            except ValueError:
                return text

        if ext == ".csv":
            with open(filepath, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    wb_no = str(row.get("运单号", "")).strip()
                    sev = str(row.get("严重程度", "")).strip()
                    cat = str(row.get("类别", "")).strip()
                    start = _normalize_time(row.get("开始时间", ""))
                    end = _normalize_time(row.get("结束时间", ""))
                    mins = _normalize_mins(row.get("时长(分钟)", ""))
                    temp = _normalize_temp(row.get("温度(℃)", ""))
                    key = f"{wb_no}|{_reverse_severity(sev)}|{_reverse_category(cat)}|{start}|{end}|{mins}|{temp}"
                    notes[key] = {
                        "处理状态": str(row.get("处理状态", "")).strip(),
                        "备注": str(row.get("备注", "")).strip(),
                        "处理人": str(row.get("处理人", "")).strip(),
                    }
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(h).strip() if h else "" for h in rows[0]]
                if "运单号" not in headers or "严重程度" not in headers:
                    continue
                for row in rows[1:]:
                    row_dict = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
                    wb_no = str(row_dict.get("运单号", "")).strip()
                    sev = str(row_dict.get("严重程度", "")).strip()
                    cat = str(row_dict.get("类别", "")).strip()
                    start = _normalize_time(row_dict.get("开始时间", ""))
                    end = _normalize_time(row_dict.get("结束时间", ""))
                    mins = _normalize_mins(row_dict.get("时长(分钟)", ""))
                    temp = _normalize_temp(row_dict.get("温度(℃)", ""))
                    key = f"{wb_no}|{_reverse_severity(sev)}|{_reverse_category(cat)}|{start}|{end}|{mins}|{temp}"
                    notes[key] = {
                        "处理状态": str(row_dict.get("处理状态", "")).strip(),
                        "备注": str(row_dict.get("备注", "")).strip(),
                        "处理人": str(row_dict.get("处理人", "")).strip(),
                    }
            wb.close()
    except Exception as e:
        click.echo(f"  警告：无法读取已有备注文件 → {e}", err=True)
    return notes


def _reverse_severity(label):
    for k, v in _SEVERITY_LABEL.items():
        if v == label:
            return k
    return label


def _reverse_category(label):
    for k, v in _CATEGORY_LABEL.items():
        if v == label:
            return k
    return label


def _write_exceptions_csv(filepath, items, note_template=False, existing_notes=None):
    _ensure_dir(filepath)
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        headers = ["序号", "严重程度", "类别", "运单号", "车牌", "设备编号",
                   "开始时间", "结束时间", "时长(分钟)", "温度(℃)", "详情"]
        if note_template:
            headers += ["处理状态", "备注", "处理人"]
        writer.writerow(headers)
        for row in _build_exception_rows(items, note_template, existing_notes):
            writer.writerow([row[h] for h in headers])


@cli.command()
@click.option("--month-dir", "-m", required=True,
              help="月份文件夹路径（含运单清单和温度子目录，支持多层）")
@click.option("--temp-zone", "-z", default=None, help="覆盖温区标准（如 -18~-15 或 2-8）")
@click.option("--output", "-o", default=None,
              help="汇总输出文件路径（XLSX，推荐）；配合 --per-customer 会输出多份文件")
@click.option("--per-customer", "-p", is_flag=True, default=False,
              help="除总表外，为每个客户单独生成一份交付文件")
@click.option("--overview-only", is_flag=True, default=False,
              help="仅显示月底总览，不导出文件")
@click.option("--merge-notes", default=None,
              help="已有备注文件路径，导出时保留原说明")
@click.option("--output-dir", "-d", default=None,
              help="指定输出目录（配合 --per-customer 使用，默认在月份目录下生成 deliverables 子目录）")
def batch(month_dir, temp_zone, output, per_customer, overview_only, merge_notes, output_dir):
    """批量处理：递归扫描月份文件夹，按客户汇总，月底总览，生成完整留痕报告"""
    if not os.path.isdir(month_dir):
        click.echo(f"错误：月份文件夹不存在 → {month_dir}", err=True)
        sys.exit(1)

    override_zone = None
    if temp_zone:
        override_zone = TempZone.parse(temp_zone)
        if not override_zone:
            click.echo(f'错误：无法解析温区标准「{temp_zone}」，请检查格式', err=True)
            sys.exit(1)
        click.echo(click.style(f"温区标准（命令行指定）：{override_zone.display()}", bold=True))

    waybill_files = find_waybill_files(month_dir)
    all_temp_files = find_temperature_files(month_dir)
    waybill_file_set = set(os.path.abspath(f) for f in waybill_files)
    temp_files = [f for f in all_temp_files
                  if os.path.abspath(f) not in waybill_file_set]

    if not waybill_files:
        click.echo(f"错误：月份文件夹中未找到运单清单文件 → {month_dir}", err=True)
        sys.exit(1)

    click.echo("正在扫描文件...")
    for wf in waybill_files:
        rel = os.path.relpath(wf, month_dir)
        click.echo(f"  运单 → {rel}")
    for tf in temp_files:
        rel = os.path.relpath(tf, month_dir)
        click.echo(f"  温度 → {rel}")

    all_waybills = []
    for wf in waybill_files:
        all_waybills.extend(load_waybills(wf))

    all_readings = load_temperature_dir(month_dir, recursive=True)

    click.echo(f"共加载 {len(all_waybills)} 条运单、{len(all_readings)} 条温度记录")

    match_result = match_waybills_readings(all_waybills, all_readings)

    customer_waybills = defaultdict(list)
    for w in all_waybills:
        cust = w.customer or "未指定客户"
        customer_waybills[cust].append(w)

    class CustomerStats:
        def __init__(self):
            self.total = 0
            self.matched = 0
            self.no_data = 0
            self.has_exception = 0
            self.max_over_temp = 0.0
            self.exceptions = []
            self.reports = []
            self.match_result = None
            self.waybills = []

    customer_stats = defaultdict(CustomerStats)
    all_reports = []
    all_exceptions = []

    for pair in match_result.matched:
        cust = pair.waybill.customer or "未指定客户"
        zone = override_zone or pair.waybill.temp_zone
        r = generate_report(pair.waybill, pair.readings, zone, pair.basis)
        all_reports.append(r)
        customer_stats[cust].reports.append(r)
        customer_stats[cust].matched += 1
        customer_stats[cust].total += 1
        customer_stats[cust].waybills.append(pair.waybill)
        if r.over_temp_minutes > customer_stats[cust].max_over_temp:
            customer_stats[cust].max_over_temp = r.over_temp_minutes
        items = detect_exceptions(pair.waybill, pair.readings, zone)
        all_exceptions.extend(items)
        customer_stats[cust].exceptions.extend(items)
        if items:
            customer_stats[cust].has_exception += 1

    for w in match_result.no_temp_data:
        cust = w.customer or "未指定客户"
        zone = override_zone or w.temp_zone
        r = generate_report(w, [], override_zone)
        all_reports.append(r)
        customer_stats[cust].reports.append(r)
        customer_stats[cust].no_data += 1
        customer_stats[cust].total += 1
        customer_stats[cust].waybills.append(w)
        items = detect_exceptions(w, [], zone)
        all_exceptions.extend(items)
        customer_stats[cust].exceptions.extend(items)
        if items:
            customer_stats[cust].has_exception += 1

    click.echo()
    click.echo("=" * 70)
    click.echo("月底总览 — 按客户风险排序")
    click.echo("=" * 70)

    ranked_customers = sorted(
        customer_stats.keys(),
        key=lambda c: (
            -customer_stats[c].has_exception / max(customer_stats[c].total, 1),
            -customer_stats[c].max_over_temp,
            -customer_stats[c].no_data,
        ),
    )

    total_wb = 0
    total_matched = 0
    total_no_data = 0
    total_exception_wb = 0

    click.echo(
        f"{'客户名称':<12} {'总票数':>6} {'有数据':>6} {'缺数据':>8} "
        f"{'有异常':>6} {'异常率':>8} {'最高超温':>10}"
    )
    click.echo("-" * 72)

    for cust in ranked_customers:
        cs = customer_stats[cust]
        total_wb += cs.total
        total_matched += cs.matched
        total_no_data += cs.no_data
        total_exception_wb += cs.has_exception

        exception_rate = cs.has_exception / cs.total * 100 if cs.total else 0
        rate_str = f"{exception_rate:.0f}%"
        max_ot_str = f"{cs.max_over_temp:.0f}分" if cs.max_over_temp > 0 else "—"

        line_parts = [f"{cust:<12} {cs.total:>6} {cs.matched:>6} {cs.no_data:>8} "
                      f"{cs.has_exception:>6}"]
        if exception_rate >= 50:
            line_parts.append(click.style(f"{rate_str:>8}", fg="red"))
        elif exception_rate >= 20:
            line_parts.append(click.style(f"{rate_str:>8}", fg="yellow"))
        else:
            line_parts.append(click.style(f"{rate_str:>8}", fg="green"))
        if cs.max_over_temp >= 30:
            line_parts.append(click.style(f"{max_ot_str:>10}", fg="red"))
        elif cs.max_over_temp >= 10:
            line_parts.append(click.style(f"{max_ot_str:>10}", fg="yellow"))
        else:
            line_parts.append(f"{max_ot_str:>10}")

        click.echo("".join(line_parts))

    click.echo("-" * 72)
    overall_rate = total_exception_wb / total_wb * 100 if total_wb else 0
    overall_max_ot = max((cs.max_over_temp for cs in customer_stats.values()), default=0)
    max_ot_str = f"{overall_max_ot:.0f}分" if overall_max_ot > 0 else "—"
    click.echo(
        f"{'合计':<12} {total_wb:>6} {total_matched:>6} {total_no_data:>8} "
        f"{total_exception_wb:>6} {overall_rate:.0f}%  {max_ot_str:>10}"
    )

    if total_no_data > 0:
        click.echo()
        click.echo(click.style("⚠ 缺数据票数排行：", fg="yellow"))
        no_data_ranked = sorted(
            customer_stats.items(),
            key=lambda x: -x[1].no_data,
        )
        for cust, cs in no_data_ranked:
            if cs.no_data > 0:
                click.echo(f"  {cust}: {cs.no_data} 票缺数据")

    if overall_max_ot > 0:
        click.echo()
        click.echo(click.style("⚠ 最高超温时长排行：", fg="yellow"))
        ot_ranked = sorted(
            customer_stats.items(),
            key=lambda x: -x[1].max_over_temp,
        )
        for cust, cs in ot_ranked:
            if cs.max_over_temp > 0:
                click.echo(f"  {cust}: {cs.max_over_temp:.0f} 分钟")

    if overview_only:
        return

    existing_notes = {}
    if merge_notes:
        click.echo()
        click.echo(f"正在合并已有备注 → {merge_notes}")
        existing_notes = _load_existing_notes(merge_notes)
        if existing_notes:
            click.echo(f"  已读取 {len(existing_notes)} 条历史备注")

    if not output and not per_customer:
        click.echo()
        click.echo("提示：使用 -o output.xlsx 可导出完整留痕报告")
        click.echo("       使用 --per-customer 可按客户生成单独交付文件")
        return

    if output:
        _ensure_dir(output)
        if not output.lower().endswith((".xlsx", ".xls")):
            output = output.rsplit(".", 1)[0] + ".xlsx"
            click.echo(f"提示：批量导出推荐 XLSX 格式，已自动调整为 → {output}")
        _export_full_excel(output, match_result, all_reports, override_zone,
                           all_exceptions=all_exceptions, existing_notes=existing_notes)
        click.echo()
        click.echo(f"总汇总表已导出 → {output}")

    if per_customer:
        if output_dir is None:
            output_dir = os.path.join(month_dir, "deliverables")
        os.makedirs(output_dir, exist_ok=True)
        click.echo()
        click.echo(f"正在为每个客户生成单独交付文件 → {output_dir}/")

        for cust in customer_stats.keys():
            cs = customer_stats[cust]
            safe_name = cust.replace("/", "_").replace("\\", "_").replace(":", "_")
            customer_file = os.path.join(output_dir, f"{safe_name}_温度留痕.xlsx")

            cust_match = type("obj", (), {})()
            cust_match.matched = []
            cust_match.no_temp_data = []
            for pair in match_result.matched:
                if pair.waybill.customer == cust:
                    cust_match.matched.append(pair)
            for w in match_result.no_temp_data:
                if w.customer == cust:
                    cust_match.no_temp_data.append(w)

            _export_full_excel(
                customer_file, cust_match, cs.reports, override_zone,
                all_exceptions=cs.exceptions,
                existing_notes=existing_notes,
                customer_name=cust,
            )
            click.echo(f"  ✓ {cust} → {os.path.basename(customer_file)}")

        click.echo(f"共 {len(customer_stats)} 份客户交付文件已生成")


def _export_full_excel(filepath, match_result, reports, override_zone,
                       all_exceptions=None, existing_notes=None, customer_name=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    title_font = Font(bold=True, size=14)

    ws1 = wb.active
    ws1.title = "匹配结果"
    if customer_name:
        ws1.cell(row=1, column=1, value=f"冷链运单温度留痕报告 — {customer_name}").font = title_font
    match_rows = _build_match_rows(match_result)
    start_row = 3 if customer_name else 1
    _write_sheet(ws1, match_rows, header_font, header_fill, header_font_white, thin_border,
                 start_row=start_row)

    ws2 = wb.create_sheet("温度摘要")
    report_rows = _build_report_rows(reports)
    _write_sheet(ws2, report_rows, header_font, header_fill, header_font_white, thin_border)

    ws3 = wb.create_sheet("关键时间点")
    kp_rows = _build_keypoint_rows(reports)
    _write_sheet(ws3, kp_rows, header_font, header_fill, header_font_white, thin_border)

    if all_exceptions is not None:
        ws4 = wb.create_sheet("异常清单")
        exc_rows = _build_exception_rows(all_exceptions, note_template=True,
                                         existing_notes=existing_notes)
        _write_sheet(ws4, exc_rows, header_font, header_fill, header_font_white, thin_border)

        headers = list(exc_rows[0].keys()) if exc_rows else []
        sev_col_idx = None
        status_col_idx = None
        for idx, h in enumerate(headers, 1):
            if h == "严重程度":
                sev_col_idx = idx
            if h == "处理状态":
                status_col_idx = idx

        if sev_col_idx:
            for row_idx in range(2, ws4.max_row + 1):
                sev_cell = ws4.cell(row=row_idx, column=sev_col_idx)
                if sev_cell.value == "严重":
                    for col_idx in range(1, ws4.max_column + 1):
                        ws4.cell(row=row_idx, column=col_idx).fill = high_fill
                elif sev_cell.value == "中等":
                    for col_idx in range(1, ws4.max_column + 1):
                        ws4.cell(row=row_idx, column=col_idx).fill = medium_fill

        if status_col_idx:
            status_col_letter = openpyxl.utils.get_column_letter(status_col_idx)
            dv = DataValidation(
                type="list",
                formula1='"已处理,待客户确认,无需处理"',
                allow_blank=True,
                showDropDown=False,
            )
            dv.error = "请从下拉列表选择"
            dv.errorTitle = "无效输入"
            dv.prompt = "请选择处理状态：已处理 / 待客户确认 / 无需处理"
            dv.promptTitle = "处理状态"
            ws4.add_data_validation(dv)
            dv.add(f"{status_col_letter}2:{status_col_letter}{ws4.max_row}")

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(filepath)


def _write_sheet(ws, rows, header_font, header_fill, header_font_white, thin_border,
                 start_row=1):
    from openpyxl.styles import Alignment
    if not rows:
        return
    headers = list(rows[0].keys())
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(h, ""))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")


def _write_excel(filepath, rows, sheet_name="Sheet1", highlight_severity=False):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    _write_sheet(ws, rows, header_font, header_fill, header_font, thin_border)

    if highlight_severity:
        headers = list(rows[0].keys()) if rows else []
        sev_col_idx = None
        status_col_idx = None
        for idx, h in enumerate(headers, 1):
            if h == "严重程度":
                sev_col_idx = idx
            if h == "处理状态":
                status_col_idx = idx

        if sev_col_idx:
            for row_idx in range(2, ws.max_row + 1):
                sev_cell = ws.cell(row=row_idx, column=sev_col_idx)
                if sev_cell.value == "严重":
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = high_fill
                elif sev_cell.value == "中等":
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = medium_fill

        if status_col_idx:
            status_col_letter = openpyxl.utils.get_column_letter(status_col_idx)
            dv = DataValidation(
                type="list",
                formula1='"已处理,待客户确认,无需处理"',
                allow_blank=True,
                showDropDown=False,
            )
            dv.error = "请从下拉列表选择"
            dv.errorTitle = "无效输入"
            dv.prompt = "请选择处理状态：已处理 / 待客户确认 / 无需处理"
            dv.promptTitle = "处理状态"
            ws.add_data_validation(dv)
            dv.add(f"{status_col_letter}2:{status_col_letter}{ws.max_row}")

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(filepath)


def _ensure_dir(filepath):
    d = os.path.dirname(filepath)
    if d:
        os.makedirs(d, exist_ok=True)


def _fmt_dt(dt):
    if dt is None:
        return "—"
    return dt.strftime("%Y-%m-%d %H:%M")


def main():
    cli()


if __name__ == "__main__":
    main()
