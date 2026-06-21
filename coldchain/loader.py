import csv
import os
from datetime import datetime
from typing import List, Optional
from .models import Waybill, TemperatureReading, TempZone


_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M",
    "%Y%m%d%H%M%S",
    "%Y%m%d%H%M",
    "%Y.%m.%d %H:%M:%S",
    "%Y.%m.%d %H:%M",
]


def _parse_datetime(text: str) -> Optional[datetime]:
    if not text or not text.strip():
        return None
    text = text.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _parse_float(text: str) -> Optional[float]:
    if not text or not text.strip():
        return None
    try:
        return float(text.strip())
    except ValueError:
        return None


_WAYBILL_COLUMN_MAP = {
    "waybill_no": ["运单号", "运单编号", "waybill_no", "waybillno", "单号"],
    "license_plate": ["车牌", "车牌号", "license_plate", "plate", "车号"],
    "device_id": ["设备编号", "设备号", "设备ID", "device_id", "deviceid", "记录仪编号"],
    "customer": ["客户", "客户名称", "customer", "客户名"],
    "temp_zone": ["温区", "温区标准", "temp_zone", "温度标准", "温区范围"],
    "departure_time": ["起运时间", "出发时间", "departure_time", "发车时间"],
    "arrival_time": ["到达时间", "到站时间", "arrival_time", "抵达时间"],
}

_TEMP_COLUMN_MAP = {
    "device_id": ["设备编号", "设备号", "设备ID", "device_id", "deviceid", "记录仪编号"],
    "timestamp": ["时间", "时间戳", "timestamp", "记录时间", "采集时间"],
    "temperature": ["温度", "温度值", "temperature", "temp", "实时温度"],
}


def _map_columns(headers: List[str], column_map: dict) -> dict:
    result = {}
    header_lower = [h.strip().lower() for h in headers]
    for target, candidates in column_map.items():
        for candidate in candidates:
            if candidate.lower() in header_lower:
                idx = header_lower.index(candidate.lower())
                result[target] = headers[idx]
                break
    return result


def load_waybills(filepath: str) -> List[Waybill]:
    waybills = []
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            col_map = _map_columns(reader.fieldnames or [], _WAYBILL_COLUMN_MAP)
            for row in reader:
                waybills.append(_build_waybill(row, col_map))
    elif ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return waybills
        headers = [str(h) if h else "" for h in rows[0]]
        col_map = _map_columns(headers, _WAYBILL_COLUMN_MAP)
        for row in rows[1:]:
            row_dict = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
            waybills.append(_build_waybill(row_dict, col_map))
        wb.close()
    else:
        raise ValueError(f"不支持的运单文件格式: {ext}")

    return waybills


def _build_waybill(row: dict, col_map: dict) -> Waybill:
    def _get(key):
        col = col_map.get(key)
        if col and col in row:
            val = row[col]
            return str(val).strip() if val else ""
        return ""

    raw_zone = _get("temp_zone")
    return Waybill(
        waybill_no=_get("waybill_no"),
        license_plate=_get("license_plate"),
        device_id=_get("device_id"),
        customer=_get("customer"),
        temp_zone=TempZone.parse(raw_zone) if raw_zone else None,
        raw_temp_zone=raw_zone,
        departure_time=_parse_datetime(_get("departure_time")),
        arrival_time=_parse_datetime(_get("arrival_time")),
    )


def load_temperature_file(filepath: str) -> List[TemperatureReading]:
    readings = []
    ext = os.path.splitext(filepath)[1].lower()
    device_from_file = os.path.splitext(os.path.basename(filepath))[0]

    if ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            col_map = _map_columns(reader.fieldnames or [], _TEMP_COLUMN_MAP)
            for row in reader:
                reading = _build_reading(row, col_map, device_from_file)
                if reading:
                    readings.append(reading)
    elif ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return readings
        headers = [str(h) if h else "" for h in rows[0]]
        col_map = _map_columns(headers, _TEMP_COLUMN_MAP)
        for row in rows[1:]:
            row_dict = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
            reading = _build_reading(row_dict, col_map, device_from_file)
            if reading:
                readings.append(reading)
        wb.close()
    else:
        raise ValueError(f"不支持的温度文件格式: {ext}")

    return readings


def _build_reading(row: dict, col_map: dict, fallback_device: str) -> Optional[TemperatureReading]:
    def _get(key):
        col = col_map.get(key)
        if col and col in row:
            val = row[col]
            return str(val).strip() if val else ""
        return ""

    device_id = _get("device_id") or fallback_device
    ts = _parse_datetime(_get("timestamp"))
    temp = _parse_float(_get("temperature"))
    if ts is None or temp is None:
        return None
    return TemperatureReading(device_id=device_id, timestamp=ts, temperature=temp)


def load_temperature_dir(dirpath: str) -> List[TemperatureReading]:
    all_readings = []
    if not os.path.isdir(dirpath):
        return all_readings
    for fname in os.listdir(dirpath):
        fpath = os.path.join(dirpath, fname)
        if os.path.isfile(fpath) and fname.lower().endswith((".csv", ".xlsx", ".xls")):
            all_readings.extend(load_temperature_file(fpath))
    all_readings.sort(key=lambda r: r.timestamp)
    return all_readings
